from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
    def tum_ogrenci_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:F4')

        top_left_cell = ws['B2']
        top_left_cell.value = " NEOS YAZILIM ÖĞRENCİ  BİLGİLERİ "

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Telefon"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Mail"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        s  =self.tablolar.tum_ogrenciler()
        for i in range(len(s)):
            
            b = "B"
            c = "C"
            d = "D"
            e = "E"
            
            z = 6 + i
            
            b = b +str(z)
            c = c +str(z)
            d = d +str(z)
            e = e +str(z)
            
            print(b)
            top_left_cell = ws[b]
            top_left_cell.value = str(s[i][0])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[c]
            top_left_cell.value = str(s[i][1])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[d]
            top_left_cell.value = str(s[i][2])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[e]
            top_left_cell.value = str(s[i][3])
            top_left_cell.font  = Font(b=True, color="00000000")
            
        
        wb.save("tum_ogrenciler.xlsx")   

    def sinif_excel(self):
        u = self.ui.comboBox_3.currentText()
        self.sertifika_excel(u)
        y = self.tablolar.sinif_bilgileri(u)
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:F4')

        top_left_cell = ws['B2']
        sinif = " NEOS YAZILIM " + u +" BİLGİLERİ "
        top_left_cell.value =  sinif

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Numara"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))

        for i in range(len(y)):
            
            b = "B"
            c = "C"
            d = "D"
            
            z = 6 + i
            
            b = b +str(z)
            c = c +str(z)
            d = d +str(z)
            
            print(b)
            top_left_cell = ws[b]
            top_left_cell.value = str(y[i][0])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[c]
            top_left_cell.value = str(y[i][1])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[d]
            top_left_cell.value = str(y[i][2])
            top_left_cell.font  = Font(b=True, color="00000000")
            
        sinif = str(sinif).replace(" ","_")
        sinif = sinif+".xlsx"
        wb.save(sinif)  
    def sertifika_excel(self,u):
        y = self.tablolar.sertifika_goster() 
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:G4')

        top_left_cell = ws['B2']
        sinif = " NEOS YAZILIM " + u + "SERTİFKASI" +" BİLGİLERİ "
        top_left_cell.value =  sinif

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "sinif"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Sertifika1"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Sertifika2"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['G5']
        top_left_cell.value = "Sertifika3"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        m = 0
        for i in range(len(y)):
            if y[i][2] == u:
                b = "B"
                c = "C"
                d = "D"
                e = "E"
                f = "F"
                g = "G"
                z = 6 + m
                
                b = b +str(z)
                c = c +str(z)
                d = d +str(z)
                e = e +str(z)
                f = f +str(z)
                g = g +str(z)
                print(b)
                top_left_cell = ws[b]
                top_left_cell.value = str(y[i][0])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[c]
                top_left_cell.value = str(y[i][1])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[d]
                top_left_cell.value = str(y[i][2])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[e]
                top_left_cell.value = str(y[i][3])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[f]
                top_left_cell.value = str(y[i][4])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[g]
                top_left_cell.value = str(y[i][5])
                top_left_cell.font  = Font(b=True, color="00000000")
                m = m+1
        sinif = str(sinif).replace(" ","_")
        sinif = sinif+".xlsx"
        wb.save(sinif)  