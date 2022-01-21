#!/usr/bin/python
from datetime import date
import datetime
from sqlite3.dbapi2 import Date
from PyQt5.QtCore import QDate
from PyQt5.QtWidgets import*
from untitled_python import*
from PyQt5.QtGui import QIntValidator , QIcon 
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

import mysql.connector





class tablo():
    def __init__(self) :
        self.baglanti  = mysql.connector.connect(
        host="sql11.freemysqlhosting.net",
        db="sql11438034",
        user=" sql11438034",
        password="AqTp9ZelPA"
        )
        self.cursor  = self.baglanti.cursor()
        
    def yonetici_tablosu(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS yonetici (İsim TEXT,sifre TEXT)")
        self.baglanti.commit()
    def yonetici_ekle(self,kullanici_adi,sifre):
        self.cursor.execute("insert into yonetici Values(%s,%s)",(kullanici_adi,sifre))
        self.baglanti.commit()
    def yonetici_kontrol(self):
        self.cursor.execute("select İsim,sifre from yonetici")
        liste = self.cursor.fetchall()
        return liste
    def ogrenci_tablosu(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS ogrenci (ogrenci_id INTEGER PRIMARY KEY ,adi_soyadi TEXT,telefon TEXT ,mail TEXT,detay TEXT ,gun INT,ay INT, yil INT)")
        self.baglanti.commit()
    def ogrenci_int_veri(self):
        self.cursor.execute("select ogrenci_id from ogrenci ")
        liste = self.cursor.fetchall()
        return liste
    def tum_ogrenciler(self):
        self.cursor.execute("select * from ogrenci ")
        liste = self.cursor.fetchall()
        return liste
    def ogrenci_ekle(self,id,a,b,c,d,e,f,g):
        self.cursor.execute("insert into ogrenci Values(%s,%s,%s,%s,%s,%s,%s,%s)",(id,a,b,c,d,e,f,g))
        self.baglanti.commit()
    def ogrenci_goster(self):
        self.cursor.execute("select adi_soyadi,telefon,mail,ogrenci_id from ogrenci ")
        liste = self.cursor.fetchall()
        return liste
    def siniflar(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS siniflar (sinif_id INTEGER PRIMARY KEY ,sinif_adi TEXT, ucreti INT,basgun INT ,basay INT ,basyil INT ,bitgun INT,bitay INT ,bityil INT)")
        self.baglanti.commit()
    def sinif_ekle(self,a,b,c,d,e,f,g,k,id):
        self.isim = a + "_"+str(c)+"_"+str(d)
        self.cursor.execute("insert into siniflar Values(%s,%s,%s,%s,%s,%s,%s,%s,%s)",(id,self.isim,b,c,d,e,f,g,k))
        self.baglanti.commit()
        self.sinif_ekleme(self.isim)
    def guncelle_ogr(self,tutar ,odenen,ogrenci):
        self.cursor.execute("update ogrenci set detay = %s where adi_soyadi = %s AND  telefon = %s  ",(tutar ,odenen,ogrenci))
        self.baglanti.commit()
    def int_veri(self):
        self.cursor.execute("select sinif_id from siniflar ")
        liste = self.cursor.fetchall()
        return liste
    def sinif_ekleme(self,gelen):
        self.sorgu = "CREATE TABLE IF NOT EXISTS "+gelen+" (ogrenci_id INTEGER PRIMARY KEY ,ogrenci_adi TEXT,numara TEXT)"
        self.cursor.execute(self.sorgu)
        self.baglanti.commit()
    def sinif_bilgileri(self,gelen):
        c = "select * from "+gelen
        self.cursor.execute(c)
        liste = self.cursor.fetchall()
        return liste
    def taksitli_odemeler(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS taksitli (ogrenci_id INT,ogrenci TEXT,sinif TEXT,ucret_talebi INT ,odenen INT ,taksit1 INT,tarih1 DATE,taksit2 INT,tarih2 DATE,taksit3 INT,tarih3 DATE)")
        self.baglanti.commit()
    def taksit_ode(self,id,isim,sinif,ucret,odenen,s1,t1,s2,t2,s3,t3):
        self.cursor.execute("insert into taksitli Values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",(id,isim,sinif,ucret,odenen,s1,t1,s2,t2,s3,t3))
        self.baglanti.commit()
    def taksit_goster(self):
        self.cursor.execute("select * from taksitli")
        liste = self.cursor.fetchall()
        return liste
    def taksilat_al1(self ,tutar ,odenen,ogrenci,sinif,tarih):
        self.cursor.execute("update taksitli set taksit1 = %s ,odenen = %s where ogrenci = %s AND  (sinif = %s AND tarih1 = %s ) ",(tutar ,odenen,ogrenci,sinif,tarih))
        self.baglanti.commit()
    def taksilat_al2(self ,tutar ,odenen,ogrenci,sinif,tarih):
        self.cursor.execute("update taksitli set taksit2 = %s ,odenen = %s where ogrenci = %s AND  (sinif = %s AND tarih2 = %s ) ",(tutar ,odenen,ogrenci,sinif,tarih))
        self.baglanti.commit()
    def taksilat_al3(self ,tutar ,odenen,ogrenci,sinif,tarih):
        self.cursor.execute("update taksitli set taksit3 = %s ,odenen = %s where ogrenci = %s AND  (sinif = %s AND tarih3 = %s ) ",(tutar ,odenen,ogrenci,sinif,tarih))
        self.baglanti.commit()
    def siniflari_goster(self):
        self.cursor.execute("select sinif_id,sinif_adi,ucreti from siniflar")
        liste = self.cursor.fetchall()
        return liste
    def sinifa_ogrenci_ekle(self,sinif,id,isim,numara):
        self.cursor.execute("insert into "+sinif+ " Values(%s,%s,%s)",(id,isim,numara))
        self.baglanti.commit()
    def pesin_odemeler(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS pesin (ogrenci_id INT,ogrenci_adi TEXT,sinif_adi TEXT,alinan_pesin INT,tarih DATE)")
        self.baglanti.commit()
    def pesin_ode(self,id,isim,sinif,alinan,tarih):
        self.cursor.execute("insert into pesin Values(%s,%s,%s,%s,%s)",(id,isim,sinif,alinan,tarih))
        self.baglanti.commit()
    def pesin_goster(self):
        self.cursor.execute("select * from pesin")
        liste = self.cursor.fetchall()
        return liste
    def sertifika(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS setifikalar (ogrenci_id INT,ogrenci_adi TEXT,sinif TEXT,sertifika1 TEXT,sertifika2 TEXT,sertifika3 TEXT)")
        self.baglanti.commit()
    def sertifika_al(self,id,isim,sinif,s1,s2,s3):
        self.cursor.execute("insert into setifikalar Values(%s,%s,%s,%s,%s,%s)",(id,isim,sinif,s1,s2,s3 ))
        self.baglanti.commit()
    def sertifika_goster(self):
        self.cursor.execute("select * from setifikalar")
        liste = self.cursor.fetchall()
        return liste
    def neden_ekle(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS nedenler (neden TEXT ,ne TEXT)")
        self.baglanti.commit()
    def neden_ogrenci(self,ekle):
        self.cursor.execute("insert into nedenler Values(%s,%s)",(ekle ,"Selam"))
        self.baglanti.commit()
    def neden_listele(self):
        self.cursor.execute("select * from nedenler")
        liste = self.cursor.fetchall()
        return liste
    def neden_sil(self,yazar):
        self.cursor.execute("DELETE FROM nedenler WHERE neden = %s",(yazar,))
        self.baglanti.commit()
    
class Neos_otomasyon(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_neos()
        self.ui.setupUi(self)
        self.giris()
        self.girisler = 0
        self.tablolar = tablo()
        self.tablolar.yonetici_tablosu()
        self.tablolar.ogrenci_tablosu()
        self.tablolar.siniflar()
        self.tablolar.pesin_odemeler()
        self.tablolar.taksitli_odemeler()
        self.tablolar.sertifika()
        self.tablolar.neden_ekle()
        self.ui.giris_buton.clicked.connect(self.goster)
        self.ui.egitim_kaydet.clicked.connect(self.sinif_ekleme)
        self.combobox_edit()
        self.ui.comboBox.activated['QString'].connect(self.combo)
        self.ui.kayit_et.clicked.connect(self.ogrenci_ekle)
        self.liste_goster()
        self.ui.listWidget.itemClicked.connect(self.tiklandi)
        self.ui.kayit_et_sinif.clicked.connect(self.kayit)
        self.ui.bul.clicked.connect(self.ogrenci_liste_goster)
        self.liste()
        self.mkc = self.ui.tableWidget.currentItem()
        print(self.mkc)
        self.ui.kayit_et_2.clicked.connect(self.elaman_ekle)
        self.listeleme()
        self.ui.listWidget_4.itemClicked.connect(self.silme)
        #self.ui.sebep_sil.clicked.connect(self.sil)
        self.taksit_goster()
        self.ui.arama.clicked.connect(self.odeme_ara)
        self.odenen = 0
        self.ui.odeme.clicked.connect(self.ode)
        self.taksit = 0
        self.odeme_al = 0
        self.ui.pushButton.clicked.connect(self.ciro)
        self.toplam = 0
        self.satis = 0
        self.comboboxlar()
        self.ui.excel_cikar.clicked.connect(self.sinif_excel)
        self.ui.kayitli_ogrenci_tablosu.clicked.connect(self.tum_ogrenci_excel)
        self.ui.silme.setText("ARA")
        self.ui.silme.clicked.connect(self.ogrenci_bilgi_goster)
        self.ui.guncelle.clicked.connect(self.guncel)
    def guncel(self):
        self.tablolar.guncelle_ogr(str(self.ui.not_arasi.toPlainText()),str(self.ui.adi_soyadi.text()),str(self.ui.telefon.text()))
        self.liste()
    def ogrenci_bilgi_goster(self):
        s = self.tablolar.tum_ogrenciler()
        m = self.ui.adi_soyadi.text()
        for i in range(len(s)):
            if m == s[i][1]:
                self.ui.kayit_tarihi.date().setDate(int(s[i][7]),int(s[i][6]),int(s[i][5]))
                self.ui.telefon.setText(str(s[i][2]))
                self.ui.mail_adrsi.setText(str(s[i][3]))
                self.ui.not_arasi.setText(str(s[i][4]))

    def tum_ogrenci_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:F4')

        top_left_cell = ws['B2']
        top_left_cell.value = " NEOS YAZILIM ÖĞRENCİ  BİLGİLERİ "

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Telefon"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Mail"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        s  =self.tablolar.tum_ogrenciler()
        for i in range(len(s)):
            
            b = "B"
            c = "C"
            d = "D"
            e = "E"
            
            z = 6 + i
            
            b = b +str(z)
            c = c +str(z)
            d = d +str(z)
            e = e +str(z)
            
            print(b)
            top_left_cell = ws[b]
            top_left_cell.value = str(s[i][0])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[c]
            top_left_cell.value = str(s[i][1])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[d]
            top_left_cell.value = str(s[i][2])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[e]
            top_left_cell.value = str(s[i][3])
            top_left_cell.font  = Font(b=True, color="00000000")
            
        
        wb.save("tum_ogrenciler.xlsx")   

    def sinif_excel(self):
        u = self.ui.comboBox_3.currentText()
        self.sertifika_excel(u)
        y = self.tablolar.sinif_bilgileri(u)
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:F4')

        top_left_cell = ws['B2']
        sinif = " NEOS YAZILIM " + u +" BİLGİLERİ "
        top_left_cell.value =  sinif

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Numara"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))

        for i in range(len(y)):
            
            b = "B"
            c = "C"
            d = "D"
            
            z = 6 + i
            
            b = b +str(z)
            c = c +str(z)
            d = d +str(z)
            
            print(b)
            top_left_cell = ws[b]
            top_left_cell.value = str(y[i][0])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[c]
            top_left_cell.value = str(y[i][1])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[d]
            top_left_cell.value = str(y[i][2])
            top_left_cell.font  = Font(b=True, color="00000000")
            
        sinif = str(sinif).replace(" ","_")
        sinif = sinif+".xlsx"
        wb.save(sinif)  
    def sertifika_excel(self,u):
        y = self.tablolar.sertifika_goster() 
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:G4')

        top_left_cell = ws['B2']
        sinif = " NEOS YAZILIM " + u + "SERTİFKASI" +" BİLGİLERİ "
        top_left_cell.value =  sinif

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "sinif"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Sertifika1"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Sertifika2"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['G5']
        top_left_cell.value = "Sertifika3"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        m = 0
        for i in range(len(y)):
            if y[i][2] == u:
                b = "B"
                c = "C"
                d = "D"
                e = "E"
                f = "F"
                g = "G"
                z = 6 + m
                
                b = b +str(z)
                c = c +str(z)
                d = d +str(z)
                e = e +str(z)
                f = f +str(z)
                g = g +str(z)
                print(b)
                top_left_cell = ws[b]
                top_left_cell.value = str(y[i][0])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[c]
                top_left_cell.value = str(y[i][1])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[d]
                top_left_cell.value = str(y[i][2])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[e]
                top_left_cell.value = str(y[i][3])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[f]
                top_left_cell.value = str(y[i][4])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[g]
                top_left_cell.value = str(y[i][5])
                top_left_cell.font  = Font(b=True, color="00000000")
                m = m+1
        sinif = str(sinif).replace(" ","_")
        sinif = sinif+".xlsx"
        wb.save(sinif)  
    def pesin_excel(self):
        
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:F4')

        top_left_cell = ws['B2']
        top_left_cell.value = " NEOS YAZILIM PESİN BİLGİLERİ "

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Sınıfı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Alınan Ücret"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        s  =self.tablolar.pesin_goster()
        for i in range(len(s)):
            
            b = "B"
            c = "C"
            d = "D"
            e = "E"
            f = "F"
            z = 6 + i
            
            b = b +str(z)
            c = c +str(z)
            d = d +str(z)
            e = e +str(z)
            f = f +str(z)
            print(b)
            top_left_cell = ws[b]
            top_left_cell.value = str(s[i][0])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[c]
            top_left_cell.value = str(s[i][1])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[d]
            top_left_cell.value = str(s[i][2])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[e]
            top_left_cell.value = str(s[i][3])
            top_left_cell.font  = Font(b=True, color="00000000")
            top_left_cell = ws[f]
            top_left_cell.value = str(s[i][4])
            top_left_cell.font  = Font(b=True, color="00000000")
        
        wb.save("Pesin.xlsx")   

    def comboboxlar(self):
        self.ui.comboBox_3.clear()
        c = self.tablolar.siniflari_goster()
        for i in c:
            self.ui.comboBox_3.addItem(i[1])
    def taksit_kalan_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:L4')

        top_left_cell = ws['B2']
        top_left_cell.value = " NEOS YAZILIM TAKSİTLERİ KALAN BİLGİLERİ "

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Sınıfı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Alınacak Ücret"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Ödenen"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['G5']
        top_left_cell.value = "İlk Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['H5']
        top_left_cell.value = "İlk Taksit Tarihi"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['I5']
        top_left_cell.value = "İkinci Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['J5']
        top_left_cell.value = "İkinci Taksit Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['K5']
        top_left_cell.value = "Üçünçü Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['L5']
        top_left_cell.value = "Üçünçü Taksit Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        s  =self.tablolar.taksit_goster()
        y = 0
        for i in range(len(s)):
            if s[i][5] != 0 or (s[i][7] != 0 or s[i][9] != 0):
                b = "B"
                c = "C"
                d = "D"
                e = "E"
                f = "F"
                g = "G"
                h = "H"
                q = "I"
                j = "J"
                k = "K"
                l = "L"
                z = 6 + y
                b = b +str(z)
                c = c +str(z)
                d = d +str(z)
                e = e +str(z)
                f = f +str(z)
                g = g +str(z)
                h = h +str(z)
                q = q +str(z)
                j = j +str(z)
                k = k +str(z)
                l = l +str(z)
                print(b)
                top_left_cell = ws[b]
                top_left_cell.value = str(s[i][0])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[c]
                top_left_cell.value = str(s[i][1])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[d]
                top_left_cell.value = str(s[i][2])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[e]
                top_left_cell.value = str(s[i][3])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[f]
                top_left_cell.value = str(s[i][4])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[g]
                top_left_cell.value = str(s[i][5])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[h]
                top_left_cell.value = str(s[i][6])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[q]
                top_left_cell.value = str(s[i][7])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[j]
                top_left_cell.value = str(s[i][8])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[k]
                top_left_cell.value = str(s[i][9])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[l]
                top_left_cell.value = str(s[i][10])
                top_left_cell.font  = Font(b=True, color="00000000")
                y = y +1
            
            wb.save("taksitleri_kalan.xlsx")
    def taksit_kalan_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B2:L4')

        top_left_cell = ws['B2']
        top_left_cell.value = " NEOS YAZILIM TAKSİTLERİ KALAN BİLGİLERİ "

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")

        top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
        top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
        top_left_cell.fill = fill = GradientFill(stop=("FC3C3C", "393E46"))
        top_left_cell.font  = Font(b=True, color="00FFFFFF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell = ws['B5']
        top_left_cell.value = "Öğrenci id"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['C5']
        top_left_cell.value = "Adı-Soyadı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['D5']
        top_left_cell.value = "Sınıfı"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['E5']
        top_left_cell.value = "Alınacak Ücret"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['F5']
        top_left_cell.value = "Ödenen"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['G5']
        top_left_cell.value = "İlk Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['H5']
        top_left_cell.value = "İlk Taksit Tarihi"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['I5']
        top_left_cell.value = "İkinci Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['J5']
        top_left_cell.value = "İkinci Taksit Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['K5']
        top_left_cell.value = "Üçünçü Taksit"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        top_left_cell = ws['L5']
        top_left_cell.value = "Üçünçü Taksit Tarih"
        top_left_cell.font  = Font(b=True, color="00000000")
        top_left_cell.fill = fill = GradientFill(stop=("F8B500", "F8B500"))
        s  =self.tablolar.taksit_goster()
        y = 0
        for i in range(len(s)):
            if s[i][5] == 0 and (s[i][7] == 0 and s[i][9] == 0):
                b = "B"
                c = "C"
                d = "D"
                e = "E"
                f = "F"
                g = "G"
                h = "H"
                q = "I"
                j = "J"
                k = "K"
                l = "L"
                z = 6 + y
                b = b +str(z)
                c = c +str(z)
                d = d +str(z)
                e = e +str(z)
                f = f +str(z)
                g = g +str(z)
                h = h +str(z)
                q = q +str(z)
                j = j +str(z)
                k = k +str(z)
                l = l +str(z)
                print(b)
                top_left_cell = ws[b]
                top_left_cell.value = str(s[i][0])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[c]
                top_left_cell.value = str(s[i][1])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[d]
                top_left_cell.value = str(s[i][2])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[e]
                top_left_cell.value = str(s[i][3])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[f]
                top_left_cell.value = str(s[i][4])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[g]
                top_left_cell.value = str(s[i][5])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[h]
                top_left_cell.value = str(s[i][6])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[q]
                top_left_cell.value = str(s[i][7])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[j]
                top_left_cell.value = str(s[i][8])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[k]
                top_left_cell.value = str(s[i][9])
                top_left_cell.font  = Font(b=True, color="00000000")
                top_left_cell = ws[l]
                top_left_cell.value = str(s[i][10])
                top_left_cell.font  = Font(b=True, color="00000000")
                y = y +1
            
            wb.save("taksitleri_Biten.xlsx")
    def ciro(self):
        pesin = self.tablolar.pesin_goster()
        tarih1 = self.ui.ciro_baslangic.text()
        tarih2 = self.ui.cirobitis.text()
        print(pesin)
        print(tarih1)
        self.taksilatlar(tarih1,tarih2)
        tarih1 = str(tarih1).split(".")
        tarih2 = str(tarih2).split(".")
        self.pesin_excel()
        self.taksit_kalan_excel()
        self.taksit_kalan_excel()
        for i in range(len(pesin)):
            s = pesin[i][4]
            s = str(s).split("-")
            if tarih1[0] <= s[0] and tarih1[2] == s[2] :
                if tarih1[1] <= s[1] :
                    if tarih2[0] >= s[0] and tarih2[2] >= s[2]:
                        print("selam")
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(pesin[i][3])
                    else:
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(pesin[i][3])
                        elif tarih2[2] > s[2]:
                            self.toplam = self.toplam + int(pesin[i][3])
            elif tarih1[0] >= s[0] and tarih1[2] == s[2] :
                if tarih1[1] <= s[1] :
                    if tarih2[0] >= s[0] and tarih2[2] >= s[2]:
                        if tarih2[1] >= s[1]: 
                            self.toplam = self.toplam + int(pesin[i][3])
                    else:
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(pesin[i][3])
                        elif tarih2[2] > s[2]:
                            self.toplam = self.toplam + int(pesin[i][3])
        self.satis = self.satis + self.toplam
        self.ui.toplam_satis.setText(str(self.satis))
        self.ui.alinacak_para.setText(str(self.toplam))
    def taksilatlar(self,tarih1,tarih2):
        self.satis = 0
        self.toplam = 0
        odeme = self.tablolar.taksit_goster()
        print(odeme)
        tarih1 = str(tarih1).split(".")
        tarih2 = self.ui.cirobitis.text()
        tarih2 = str(tarih2).split(".")
        for i in range(len(odeme)):
            print(odeme[i][6])
            s = odeme[i][6] 
            s = str(s).split("-")
            print(s)
            if tarih1[0] <= s[0] and tarih1[2] == s[2] :
                if tarih1[1] <= s[1] :
                    if tarih2[0] >= s[0] and tarih2[2] >= s[2]:
                        print("selam")
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
                    else:
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
                        elif tarih2[2] > s[2]:
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
            elif tarih1[0] >= s[0] and tarih1[2] == s[2] :
                if tarih1[1] <= s[1] :
                    if tarih2[0] >= s[0] and tarih2[2] >= s[2]:
                        if tarih2[1] >= s[1]: 
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
                    else:
                        if tarih2[1] >= s[1] and tarih2[2] >= s[2]: 
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
                        elif tarih2[2] > s[2]:
                            self.toplam = self.toplam + int(odeme[i][4])
                            self.satis = self.satis+int(odeme[i][5])+int(odeme[i][7])+int(odeme[i][9])
            
    def ode (self):
        ogr = self.ui.adi_soyadi_2.text()
        sinif = self.ui.sinif.text()
        tutar = self.ui.odeme_tutari.text()
        try:
            self.odenen = self.odenen + int(tutar)
        except ValueError:
            self.ui.statusbar.showMessage("Taksit Kısmı Boş::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        tarih = self.ui.kayit_tarihi_2.text()
        if self.taksit == 1 and self.odeme_al == 1:
            self.tablolar.taksilat_al1(0,self.odenen,ogr,sinif,tarih)
            self.taksit_goster()
            self.odeme_al = 0
            self.ui.statusbar.showMessage("Ödeme Başarılı Hayırlı Olsun ::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        elif self.taksit == 2 and self.odeme_al == 1:
            self.tablolar.taksilat_al2(0,self.odenen,ogr,sinif,tarih)
            self.taksit_goster()
            self.odeme_al = 0
            self.ui.statusbar.showMessage("Ödeme Başarılı Hayırlı Olsun ::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        elif self.taksit == 3 and self.odeme_al == 1:
            self.tablolar.taksilat_al3(0,self.odenen,ogr,sinif,tarih)
            self.taksit_goster()
            self.odeme_al = 0
            self.ui.statusbar.showMessage("Ödeme Başarılı Hayırlı Olsun ::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        else:
            self.ui.statusbar.showMessage("Taksit bulunamadı ::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
    def odeme_ara(self):
        adi = self.ui.adi_soyadi_2.text()
        ogr = self.tablolar.taksit_goster()
        print("Çalıitı")
        for i in range(len(ogr)):
            
            if adi == ogr[i][1] and ogr[i][2] == self.ui.sinif.text() :
                if ogr[i][5] != 0:
                    self.ui.odeme_tutari.setText(str(ogr[i][5]))
                    s = ogr[i][3] - ogr[i][4]
                    self.ui.label_42.setText(str(s))
                    self.odenen = ogr[i][4]
                    t = ogr[i][6]
                    t = str(t)
                    t = t.split("-")
                    tarih = QDate(int(t[2]),int(t[1]),int(t[0]))
                    self.ui.kayit_tarihi_2.setDate(tarih)
                    self.taksit = 1
                    self.odeme_al = 1
                    
                else:
                    if ogr[i][7] != 0:
                        self.ui.odeme_tutari.setText(str(ogr[i][7]))
                        s = ogr[i][3] - ogr[i][4]
                        self.ui.label_42.setText(str(s))
                        t = ogr[i][8]
                        self.odenen = ogr[i][4]
                        t = str(t)
                        t = t.split("-")
                        tarih = QDate(int(t[2]),int(t[1]),int(t[0]))
                        self.ui.kayit_tarihi_2.setDate(tarih)
                        self.taksit = 2
                        self.odeme_al = 1
                    else:
                        if ogr[i][9] != 0:
                            self.ui.odeme_tutari.setText(str(ogr[i][9]))
                            s = ogr[i][3] - ogr[i][4]
                            self.ui.label_42.setText(str(s))
                            t = ogr[i][10]
                            self.odenen = ogr[i][4]
                            t = str(t)
                            t = t.split("-")
                            tarih = QDate(int(t[2]),int(t[1]),int(t[0]))
                            self.ui.kayit_tarihi_2.setDate(tarih)
                            self.taksit = 3
                            self.odeme_al = 1
                        else:

                            self.odeme_al = 0
        
    def taksit_goster(self):
        ogr = self.tablolar.taksit_goster()
        print(len(ogr))
        self.ui.tableWidget_2.setRowCount(len(ogr))
        #self.ui.tableWidget.insertRow(len(ogr))
        for i in range(len(ogr)):
            for j in range(11):
                
                c = str(ogr[i][j])
                        
                self.ui.tableWidget_2.setItem(i,j, QTableWidgetItem(c))
                
    
    def silme(self):
        j =self.ui.listWidget_4.currentItem()
        u = self.tablolar.neden_listele()
        for i in u :
            if j.text() == i[0]:
                self.ui.kayit_olamama_nedeni.setText(j.text())
    def sil(self):
        self.tablolar.neden_sil(self.ui.kayit_olamama_nedeni.text)
        self.listeleme()
    def gel(self):
        print(self.mkc)
        print("Selam")
    def elaman_ekle(self):
        ekle = self.ui.kayit_olamama_nedeni.text()
        if ekle == "":
            pass
        else:
            self.tablolar.neden_ogrenci(ekle)
            self.listeleme()
    def listeleme(self):
        
        self.ui.listWidget_4.clear()
        
        u = self.tablolar.neden_listele()
        for i in u :
            self.ui.listWidget_4.addItem(i[0])
       

    def liste(self):
        ogr = self.tablolar.tum_ogrenciler()
        print(len(ogr))
        self.ui.tableWidget.setRowCount(len(ogr))
        #self.ui.tableWidget.insertRow(len(ogr))
        for i in range(len(ogr)):
            for j in range(6):
                if j == 5 :
                    c = str(ogr[i][5])+ "."+str(ogr[i][6])+"."+str(ogr[i][7])
                        
                    self.ui.tableWidget.setItem(i,j, QTableWidgetItem(c))
                else:
                    c = str(ogr[i][j])
                    self.ui.tableWidget.setItem(i,j, QTableWidgetItem(c))
    def sinif_kayit(self,sinif,id,isim,numara):
        self.tablolar.sinifa_ogrenci_ekle(sinif,id,isim,numara)
        self.ui.statusbar.showMessage("Kayıt Edildi ::))",3000)
        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
    def pesin(self,id,isim,sinif,alinan,tarih):
        self.tablolar.pesin_ode(id,isim,sinif,alinan,tarih)
    def kayit(self):
        if self.ui.ogrenci_adi.text() != "Öğrenci Soyadı":
            if self.ui.label_19.text() != "Eğitim Seçmelisiniz !!!":
                u = self.ui.comboBox_2.currentIndex()
                y = self.ui.comboBox.currentText()
                print(y)
                print(u)
                if u == 1:
                    if self.ui.pesin.text() == "":
                        self.ui.statusbar.showMessage("Ücreti Doldurunuz ::))",3000)
                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")    
                    else:
                        try:
                            y = self.ui.comboBox.currentText()
                            k = int(self.ui.pesin.text())
                            self.sinif_kayit(y,self.id,self.isim,self.numara)
                            tarih = self.ui.dateEdit_3.text()
                            self.pesin(self.id,self.isim,y,k,tarih)
                            self.sertifika(self.id,self.isim,y)
                            
                            self.ui.statusbar.showMessage("Kayıt Başarılı ::))",3000)
                            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")  
                        except:
                            self.ui.statusbar.showMessage("Ücreti Yanlış Doldurunuz ::))",3000)
                            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                elif u == 0:
                    self.ui.statusbar.showMessage("Ödeme Yöntemini Doldurunuz ::))",3000)
                    self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                else:
                    try:
                        b = int(self.ui.onodeme.text())
                        
                        if self.ui.onodeme.text() != "":
                            s = self.ui.taksit_combo.currentIndex()
                            if s == 0 :
                                self.ui.statusbar.showMessage("Taksit Seçiniz ::))",3000)
                                self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                            elif s ==1:
                                try:
                                    b = int(self.ui.ilk_taksit.text())
                                    if self.ui.ilk_taksit.text() != "":
                                        y = self.ui.comboBox.currentText()
                                        ucret = int(self.ui.ucret.text())
                                        on_odeme = int(self.ui.onodeme.text())
                                        ilk_taksit = int(self.ui.ilk_taksit.text())
                                        tarih = self.ui.dateEdit_2.text()
                                        self.tablolar.taksit_ode(self.id,self.isim,y,ucret,on_odeme,ilk_taksit,tarih,0,tarih,0,tarih)
                                        self.sertifika(self.id,self.isim,y)
                                        self.ui.statusbar.showMessage("Kayıt Edildi  ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                    else:
                                        self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                except ValueError:
                                    self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                    self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                
                            elif s == 2:
                                try:
                                    b = int(self.ui.ilk_taksitt.text())
                                    a = int(self.ui.ilk_taksitt_2.text())
                                    if a == "" or b == "":
                                        
                                        self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                    else:
                                        y = self.ui.comboBox.currentText()
                                        ucret = int(self.ui.ucret.text())
                                        on_odeme = int(self.ui.onodeme.text())
                                        t1 = self.ui.ilk_taksit_tarih.text()
                                        t2 = self.ui.ikinci_taksit_tarih.text()
                                        self.tablolar.taksit_ode(self.id,self.isim,y,ucret,on_odeme,b,t1,a,t2,0,t2)
                                        self.sertifika(self.id,self.isim,y)
                                        self.ui.statusbar.showMessage("Kayıt Edildi  ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                except ValueError:
                                    self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                    self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                
                                
                            else:
                                try:
                                    a = int(self.ui.taksit_3_ilk.text())
                                    b = int(self.ui.taksit_3_ilk_2.text())
                                    c = int(self.ui.taksit_3_ilk_3.text())
                                    if a == "" or b == "" or c == "":
                                        self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                    else:
                                        y = self.ui.comboBox.currentText()
                                        ucret = int(self.ui.ucret.text())
                                        on_odeme = int(self.ui.onodeme.text())
                                        t1 = self.ui.ilk_3_taksit.text()
                                        t2 = self.ui.ilk_3_taksit_2.text()
                                        t3 = self.ui.ilk_3_taksit_3.text()
                                        self.tablolar.taksit_ode(self.id,self.isim,y,ucret,on_odeme,a,t1,b,t2,c,t3)
                                        self.sertifika(self.id,self.isim,y)
                                        self.ui.statusbar.showMessage("Kayıt Edildi  ::))",3000)
                                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                except ValueError:
                                    self.ui.statusbar.showMessage("Taksitleri Düzgün Doldurunuz ::))",3000)
                                    self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                                
                                
                        else:
                            self.ui.statusbar.showMessage("Ön Ödemeyi Doldurunuz ::))",3000)
                            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                    except ValueError:
                        self.ui.statusbar.showMessage("Ön Ödemeyi Doldurunuz ::))",3000)
                        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
            else:
                self.ui.statusbar.showMessage("Sınıfı  Doldurunuz ::))",3000)
                self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")  
        else:
            self.ui.statusbar.showMessage("Öğrenci Seçin  ::))",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
    def sertifika(self,ogrenci_id,ogrenci_adi,sinif):
        iccw = self.ui.checkBox.isChecked()
        katilim = self.ui.checkBox_2.isChecked()
        universite = self.ui.checkBox_3.isChecked()
        i = ""
        k = ""
        u = ""
        if iccw == 1:
            i = "ICCW"
        if katilim == 1:
            k = "KATILIM"
        if universite == 1:
            u = "ÜNİVERSİTE"
        self.tablolar.sertifika_al(ogrenci_id,ogrenci_adi,sinif,i,k,u)
            
    def tiklandi(self):
        j =self.ui.listWidget.currentItem()
        u = self.tablolar.ogrenci_goster()
        for i in u :
            if j.text() == i[0]:
                self.ui.label_16.setText(i[1])
                self.ui.label_18.setText(i[2]) 
                self.ui.ogrenci_adi.setText(j.text())
                self.id = i[3]
                self.isim = i[0]
                self.numara = i[1] 
                
    def liste_goster(self):
        self.ui.listWidget.clear()
        self.ui.listWidget_2.clear()
        u = self.tablolar.ogrenci_goster()
        for i in u :
            self.ui.listWidget.addItem(i[0])
        for i in u :
            self.ui.listWidget_2.addItem(i[1])
    def ogrenci_liste_goster(self):
        u = self.tablolar.ogrenci_goster()
        if self.ui.ogrenci_adi_2.text() != "":
            self.ui.listWidget.clear()
            self.ui.listWidget_2.clear()
            for i in u:
                if self.ui.ogrenci_adi_2.text() == i[0]:
                    self.ui.listWidget.addItem(i[0])
                    self.ui.listWidget_2.addItem(i[1])
                
        else:
            self.liste_goster()
    def odeme_tablosu(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(10)
    def combo(self):
        d = self.ui.comboBox.currentIndex()
        c = self.tablolar.siniflari_goster()
        m = c[d][2]
        self.ui.ucret.setText(str(m))    
    def giris(self):
        self.ui.stackedWidget.setCurrentIndex(0)
    def kayit_sayfasi(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(1)
    def ogrenci_bilgileri_sayfasi(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(2)
    def sinifa_ogrenci_ekle(self):
        if self.girisler == 1:    
            self.ui.stackedWidget.setCurrentIndex(3)
            
    def sinif_olustur(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(4)
    def uye_olusturma(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(6) 
    def ciro_tablosu(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(8)
    def ogrenciler_tablosu(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(9)
    def kayit_olmam_nedeni(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(7)
    def tablo(self):
        if self.girisler == 1:
            self.ui.stackedWidget.setCurrentIndex(5)
    def goster(self):
        a = self.ui.kullanici_edit.text()
        b = self.ui.sifre_edit.text()
        liste = self.tablolar.yonetici_kontrol()
        for i in liste:
            if a == i[0] and b == i[1]:
                self.girisler = 1
                self.calistir()
                self.ui.statusbar.showMessage("Giriş Başarılı ::))",3000)
                self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                self.ui.kullanici_edit.setText("")
                self.ui.sifre_edit.setText("")
                self.kayit_sayfasi()

            else:
                self.ui.statusbar.showMessage("Kullanıcı Bulunamadı !!!",3000)
                self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
                self.ui.kullanici_edit.setText("")
                self.ui.sifre_edit.setText("")
    def sinif_ekleme(self):
        a = self.ui.sinif_adi.text()
        
        try:
            b = int(self.ui.egitim_tutari.text())
            c = self.ui.egitim_tarihi.date().day()
            d = self.ui.egitim_tarihi.date().month()
            e = self.ui.egitim_tarihi.date().year()
            f = self.ui.egitim_bitis_tarihi.date().day()
            g = self.ui.egitim_bitis_tarihi.date().month()
            k = self.ui.egitim_bitis_tarihi.date().year()
            id = self.tablolar.int_veri()
            buyuk = 0
            print(id)
            for i in id:
                if i[0] > buyuk:
                    buyuk = i[0]
            buyuk = buyuk + 1
            self.tablolar.sinif_ekle(a,b,c,d,e,f,g,k,buyuk)
            self.combobox_edit()
        except ValueError:
            self.ui.statusbar.showMessage("Düzgün Doldurunuz  !!!",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
            self.ui.egitim_tutari.setText("")
    def ogrenci_ekle(self):
        id = self.tablolar.ogrenci_int_veri()
        buyuk = 0
        print(id)
        for i in id:
            if i[0] > buyuk:
                buyuk = i[0]
            buyuk = buyuk + 1
        a = self.ui.lineEdit.text() 
        b = self.ui.dateEdit.date().day()
        c = self.ui.dateEdit.date().month()
        d = self.ui.dateEdit.date().year()
        e = self.ui.lineEdit_3.text()
        f = self.ui.lineEdit_2.text()
        g = self.ui.not_adress_kayit.toPlainText()
        self.tablolar.ogrenci_ekle(buyuk,a,e,f,g,b,c,d)
        self.ui.statusbar.showMessage("Kayıt  Başarılı ::))",3000)
        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        self.liste_goster()    
    def combobox_edit(self):
        self.ui.comboBox.clear()
        c = self.tablolar.siniflari_goster()
        for i in c:
            self.ui.comboBox.addItem(i[1])
        

    def cik(self):
        self.girisler = 0
        self.ui.statusbar.showMessage("Çıkış Başarılı ::))",3000)
        self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
        self.giris()
    def calistir(self):
        if self.girisler == 1: 
            self.ui.action_k.triggered.connect(self.cik)
            self.ui.action_renci_Kaydet.triggered.connect(self.kayit_sayfasi)
            self.ui.action_renci_Bilgileri.triggered.connect(self.ogrenci_bilgileri_sayfasi)
            self.ui.actionS_n_fa_renci_Ekle.triggered.connect(self.sinifa_ogrenci_ekle)
            self.ui.actionS_n_f_Olu_tur.triggered.connect(self.sinif_olustur)
            self.ui.actionNeos_Yaz_l_m_ye_olu_tur.triggered.connect(self.uye_olusturma)
            self.ui.actionCiro_Tablosu.triggered.connect(self.ciro_tablosu)
            self.ui.actionDurum_Tablosu.triggered.connect(self.ogrenciler_tablosu)
            self.ui.actionKay_t_Olamama_Nedenleri.triggered.connect(self.kayit_olmam_nedeni)
            self.ui.kayit_ol.clicked.connect(self.uye_kayit)
            self.ui.actionExcel_Tablosu.triggered.connect(self.tablo)
            self.ui.action_deme_Al.triggered.connect(self.odeme_tablosu)
        
    def uye_kayit(self):

        if self.ui.parola.text() == self.ui.parola_2.text():
            self.tablolar.yonetici_ekle(self.ui.parola.text(),self.ui.parola_2.text())
            self.ui.statusbar.showMessage("Kayıt Oldunu ;=}",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
            self.ui.parola.setText("")
            self.ui.parola_2.setText("")
            self.ui.kullanici_adi.setText("")
            
            
        else: 
            self.ui.statusbar.showMessage("Şifreler Uyuşmuyor !!!",3000)
            self.ui.statusbar.setStyleSheet("color: rgb(0, 0, 0); font-weight:bold;")
            self.ui.parola.setText("")
            self.ui.parola_2.setText("")
            self.goster()

app = QApplication([])
window =Neos_otomasyon()
window.show()
app.exec_()

